from openpyxl import load_workbook

wb=load_workbook('c.xlsx')
ws=wb.active

ws['A11']="=SUM(A1:A10)"

ws.append([11,22,33,44,55,66,77,88,99]) #한줄삽입

wb.save('c.xlsx')
wb.close()