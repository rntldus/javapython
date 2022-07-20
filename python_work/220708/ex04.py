#엑셀 만들기

from openpyxl import Workbook

wb=Workbook()
ws=wb.active

ws.cell(1,1,'123123')
# ws['A1']='123123'
ws['A2']='안녕하세요'
ws.title="처음에자동sheet"
ws.sheet_properties.tabcolor="1072BA"

ms1=wb.create_sheet("MySheet")
ms1['B1']='하이하이'
ms1['B2']='하하하,,,'

ms2=wb.create_sheet("MySheet",0)
ms2['C1']='test'
ms2['C2']='test'    

ms3=wb["MySheet"]

wb.save('a.xlsx')
wb.close()
